import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('Softball AND Baseball Banner & Sponsorship Log.xlsx')
master_ws = wb["Master Sponsor List"]

# Prepare comprehensive data for website
all_sponsors = []

print("Extracting all sponsor data for transparent website display...")

for row_idx in range(2, master_ws.max_row + 1):
    sponsor_type = master_ws.cell(row=row_idx, column=1).value
    company = master_ws.cell(row=row_idx, column=2).value
    
    # Skip empty rows
    if not company or company == "":
        continue
    
    contact = master_ws.cell(row=row_idx, column=3).value or ""
    phone = master_ws.cell(row=row_idx, column=4).value or ""
    email = master_ws.cell(row=row_idx, column=5).value or ""
    address = master_ws.cell(row=row_idx, column=6).value or ""
    division = master_ws.cell(row=row_idx, column=7).value or ""
    status = master_ws.cell(row=row_idx, column=8).value or ""
    year_2025 = master_ws.cell(row=row_idx, column=9).value or ""
    year_2024 = master_ws.cell(row=row_idx, column=10).value or ""
    year_2023 = master_ws.cell(row=row_idx, column=11).value or ""
    year_2022 = master_ws.cell(row=row_idx, column=12).value or ""
    year_2021 = master_ws.cell(row=row_idx, column=13).value or ""
    year_2020 = master_ws.cell(row=row_idx, column=14).value or ""
    notes = master_ws.cell(row=row_idx, column=15).value or ""
    
    sponsor_data = {
        "id": row_idx - 1,
        "sponsorType": str(sponsor_type),
        "company": str(company),
        "contact": str(contact),
        "phone": str(phone),
        "email": str(email),
        "address": str(address),
        "division": str(division),
        "status": str(status),
        "year2025": str(year_2025) if year_2025 else "",
        "year2024": str(year_2024) if year_2024 else "",
        "year2023": str(year_2023) if year_2023 else "",
        "year2022": str(year_2022) if year_2022 else "",
        "year2021": str(year_2021) if year_2021 else "",
        "year2020": str(year_2020) if year_2020 else "",
        "notes": str(notes)
    }
    
    all_sponsors.append(sponsor_data)

# Export to JSON for website
json_filename = "sponsors_data.json"
with open(json_filename, 'w', encoding='utf-8') as f:
    json.dump(all_sponsors, f, indent=2, ensure_ascii=False)

print(f"\n✓ Exported {len(all_sponsors)} sponsors to {json_filename}")
print(f"✓ All columns included for full transparency")

# Summary
print(f"\nSummary:")
print(f"  Total Sponsors: {len(all_sponsors)}")
baseball = sum(1 for s in all_sponsors if s['division'] == 'Baseball')
softball = sum(1 for s in all_sponsors if s['division'] == 'Softball')
print(f"  Baseball: {baseball}")
print(f"  Softball: {softball}")

active_2025 = sum(1 for s in all_sponsors if s['year2025'] and s['year2025'] != '')
print(f"  Active 2025: {active_2025}")

wb.close()
print(f"\n✓ Ready for transparent website display!")
