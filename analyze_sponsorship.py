import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Load the workbook
wb = openpyxl.load_workbook('Softball AND Baseball Banner & Sponsorship Log.xlsx')

print("=" * 80)
print("SPONSORSHIP FILE ANALYSIS")
print("=" * 80)

print(f"\nSheet names: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns\n")
    
    # Print first 20 rows to understand structure
    print("First 20 rows:")
    print("-" * 80)
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                row_data.append(f"Col{col_idx}: {value}")
        if row_data:
            print(f"Row {row_idx}: {' | '.join(row_data)}")
    print()

wb.close()
