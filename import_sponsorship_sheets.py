import math
from datetime import date, datetime

import openpyxl
from sqlmodel import Session, select

from bucksport_api.database import engine
from bucksport_api.models import SponsorshipSheetMeta, SponsorshipSheetRow


EXCEL_FILE = "Softball AND Baseball Banner & Sponsorship Log.xlsx"
SHEETS = [
    "Master Sponsor List",
    "Softball Banners - Current",
    "Softball Banners - Team Sponsor",
    "Baseball Banners - Current",
]


def _json_safe(value):
    if value is None:
        return ""

    # float nan
    if isinstance(value, float) and math.isnan(value):
        return ""

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    return value


def import_sheet(session: Session, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]

    # Preserve exact column headers, including multiline strings.
    header_row = ws[1]

    # Some spreadsheets have trailing empty columns; trim to last non-empty header
    last_col = 0
    for i, cell in enumerate(header_row, start=1):
        if cell.value not in (None, ""):
            last_col = i

    if last_col == 0:
        return 0

    headers = [header_row[i - 1].value for i in range(1, last_col + 1)]
    columns = [str(h) if h is not None else "" for h in headers]

    meta = session.get(SponsorshipSheetMeta, sheet_name)
    if not meta:
        meta = SponsorshipSheetMeta(sheet_name=sheet_name, columns=columns)
    else:
        meta.columns = columns
        meta.updated_at = datetime.utcnow()

    session.add(meta)
    session.commit()

    # Replace existing rows for this sheet
    existing_rows = session.exec(
        select(SponsorshipSheetRow).where(SponsorshipSheetRow.sheet_name == sheet_name)
    ).all()
    for r in existing_rows:
        session.delete(r)
    session.commit()

    inserted = 0
    for row_num in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_num, column=col_num).value for col_num in range(1, last_col + 1)]
        row_data = {columns[i]: _json_safe(values[i]) for i in range(len(columns))}

        # Skip completely empty rows
        if all(v in ("", None) for v in row_data.values()):
            continue

        session.add(
            SponsorshipSheetRow(
                sheet_name=sheet_name,
                row_index=row_num,
                data=row_data,
                updated_at=datetime.utcnow(),
            )
        )
        inserted += 1

    session.commit()
    return inserted


def main() -> None:
    with Session(engine) as session:
        total = 0
        for sheet in SHEETS:
            inserted = import_sheet(session, sheet)
            print(f"✅ Imported {inserted} rows for sheet: {sheet}")
            total += inserted

        print(f"\n✅ Done. Total rows imported: {total}")


if __name__ == "__main__":
    main()
