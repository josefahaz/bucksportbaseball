"""FastAPI backend for Bucksport Youth Softball/Baseball program."""
from datetime import datetime
from typing import List, Optional
import logging
import os
from pathlib import Path

from contextlib import asynccontextmanager
from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlmodel import Session, select
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
import openpyxl
from io import BytesIO

from database import get_session, init_db
from models import Event, Player, PlayerBase, Team, InventoryItem, BoardMember, Coach, Location, ScheduleEvent
from auth_routes import router as auth_router
from auth import get_current_user, get_current_fundraising_editor, can_edit_fundraising
from auth_models import User
from seed_users import seed_users
from seed_inventory import seed_inventory
from seed_board_coaches import seed_all as seed_board_coaches
from update_inventory_divisions import update_divisions

# Load environment-specific .env file
# Set ENVIRONMENT=production on production server
environment = os.getenv('ENVIRONMENT', 'development')
if environment == 'production':
    load_dotenv('.env.production')
else:
    load_dotenv('.env')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.info(f"Running in {environment} mode")

@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        logger.info("Initializing database...")
        init_db()
        logger.info("Database initialized successfully")
        
        # Seed data on startup (will skip if already seeded)
        logger.info("Seeding users...")
        seed_users()
        logger.info("Seeding inventory...")
        seed_inventory()
        logger.info("Seeding board members and coaches...")
        seed_board_coaches()
        logger.info("Updating inventory divisions...")
        update_divisions()
        logger.info("Startup complete!")
    except Exception as e:
        logger.error(f"Error during startup: {e}")
        # Don't re-raise - allow app to start even if seeding fails
    yield

# Create the FastAPI app
app = FastAPI(
    title="Bucksport Baseball/Softball API",
    version="0.2.0",
    lifespan=lifespan,
)

# Apply CORS middleware to allow all origins
origins = [
    "https://admin.bucksportll.org",
    "https://bucksportll.org",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:5500",
    "http://127.0.0.1:5500",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global exception handler to ensure CORS headers are sent on errors
from fastapi.responses import JSONResponse
from fastapi import Request

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled exception: {exc}")
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc)},
        headers={
            "Access-Control-Allow-Origin": request.headers.get("origin", "*"),
            "Access-Control-Allow-Credentials": "true",
        }
    )

# Include routers
app.include_router(auth_router)

# Resolve static directory path
BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR.parent / "local_site"

# Health check endpoint (for monitoring and uptime checks)
@app.get("/api/health")
def health_check():
    """Health check endpoint for monitoring service status."""
    return {"status": "ok", "message": "Server is running"}

# Serve the main page
@app.get("/", response_class=FileResponse, include_in_schema=False)
async def read_index():
    return str(STATIC_DIR / "index.html")

# ----------------- Team endpoints -----------------
@app.post("/api/teams", response_model=Team, status_code=status.HTTP_201_CREATED)
def create_team(team: Team, session: Session = Depends(get_session)):
    session.add(team)
    session.commit()
    session.refresh(team)
    return team

@app.get("/api/teams", response_model=List[Team])
def read_teams(session: Session = Depends(get_session)):
    return session.exec(select(Team)).all()

# ----------------- Player endpoints -----------------
@app.post("/api/players", response_model=Player, status_code=status.HTTP_201_CREATED)
def register_player(player_data: PlayerBase, session: Session = Depends(get_session)):
    try:
        logger.info(f"Received player registration: {player_data.dict()}")
        player = Player.from_orm(player_data)
        session.add(player)
        session.commit()
        session.refresh(player)
        logger.info(f"Player registered successfully: {player.id}")
        return player
    except Exception as e:
        session.rollback()
        logger.error(f"Error registering player: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

@app.get("/api/players", response_model=List[Player])
def read_players(session: Session = Depends(get_session)):
    return session.exec(select(Player)).all()

@app.get("/api/players/{player_id}", response_model=Player)
def read_player(player_id: int, session: Session = Depends(get_session)):
    player = session.get(Player, player_id)
    if not player:
        raise HTTPException(status_code=404, detail="Player not found")
    return player

# ----------------- Event endpoints -----------------
@app.post("/api/events", response_model=Event, status_code=status.HTTP_201_CREATED)
def create_event(event: Event, session: Session = Depends(get_session)):
    if event.end_time <= event.start_time:
        raise HTTPException(status_code=400, detail="end_time must be after start_time")
    session.add(event)
    session.commit()
    session.refresh(event)
    return event

@app.get("/api/events", response_model=List[Event])
def read_events(team_id: int | None = None, session: Session = Depends(get_session)):
    query = select(Event)
    if team_id is not None:
        query = query.where(Event.team_id == team_id)
    return session.exec(query.order_by(Event.start_time)).all()

# ----------------- Board Members endpoints (DATABASE) -----------------
@app.get("/api/board-members")
def get_board_members(session: Session = Depends(get_session)):
    """Get all board members from database."""
    statement = select(BoardMember)
    members = session.exec(statement).all()
    return [
        {
            "id": m.id,
            "name": m.name,
            "position": m.position,
            "division": m.division,
            "email": m.email,
            "phone": m.phone
        }
        for m in members
    ]

@app.put("/api/board-members/{member_id}")
def update_board_member(member_id: int, member_data: dict, session: Session = Depends(get_session)):
    """Update a board member in the database."""
    member = session.get(BoardMember, member_id)
    if not member:
        raise HTTPException(status_code=404, detail="Board member not found")
    
    # Update fields
    if "name" in member_data:
        member.name = member_data["name"]
    if "position" in member_data:
        member.position = member_data["position"]
    if "division" in member_data:
        member.division = member_data["division"]
    if "email" in member_data:
        member.email = member_data["email"]
    if "phone" in member_data:
        member.phone = member_data["phone"]
    
    member.updated_at = datetime.utcnow()
    session.add(member)
    session.commit()
    session.refresh(member)
    
    logger.info(f"Updated board member {member_id}: {member.name}")
    
    return {
        "id": member.id,
        "name": member.name,
        "position": member.position,
        "division": member.division,
        "email": member.email,
        "phone": member.phone
    }

# ----------------- Coaches endpoints (DATABASE) -----------------
@app.get("/api/coaches")
def get_coaches(session: Session = Depends(get_session)):
    """Get all coaches from database."""
    statement = select(Coach)
    coaches = session.exec(statement).all()
    return [
        {
            "id": c.id,
            "name": c.name,
            "email": c.email,
            "phone": c.phone,
            "team_name": c.team_name,
            "division": c.division
        }
        for c in coaches
    ]

@app.put("/api/coaches/{coach_id}")
def update_coach(coach_id: int, coach_data: dict, session: Session = Depends(get_session)):
    """Update a coach in the database."""
    coach = session.get(Coach, coach_id)
    if not coach:
        raise HTTPException(status_code=404, detail="Coach not found")
    
    # Update fields
    if "name" in coach_data:
        coach.name = coach_data["name"]
    if "email" in coach_data:
        coach.email = coach_data["email"]
    if "phone" in coach_data:
        coach.phone = coach_data["phone"]
    if "team_name" in coach_data:
        coach.team_name = coach_data["team_name"]
    if "division" in coach_data:
        coach.division = coach_data["division"]
    
    coach.updated_at = datetime.utcnow()
    session.add(coach)
    session.commit()
    session.refresh(coach)
    
    logger.info(f"Updated coach {coach_id}: {coach.name}")
    
    return {
        "id": coach.id,
        "name": coach.name,
        "email": coach.email,
        "phone": coach.phone,
        "team_name": coach.team_name,
        "division": coach.division
    }

@app.post("/api/coaches")
def create_coach(coach_data: dict, session: Session = Depends(get_session)):
    """Create a new coach."""
    coach = Coach(
        name=coach_data.get("name", ""),
        email=coach_data.get("email", "N/A"),
        phone=coach_data.get("phone", "N/A"),
        team_name=coach_data.get("team_name"),
        division=coach_data.get("division")
    )
    session.add(coach)
    session.commit()
    session.refresh(coach)
    
    logger.info(f"Created coach: {coach.name}")
    
    return {
        "id": coach.id,
        "name": coach.name,
        "email": coach.email,
        "phone": coach.phone,
        "team_name": coach.team_name,
        "division": coach.division
    }

# ----------------- Schedule endpoints (DATABASE) -----------------
@app.get("/api/schedule")
def get_schedule(session: Session = Depends(get_session)):
    """Get all scheduled events from database."""
    try:
        statement = select(ScheduleEvent)
        events = session.exec(statement).all()
        
        # If no events in DB, return sample data for now
        if not events:
            return [
                {"id": 1, "title": "Opening Day", "date": "2025-04-05", "time": "10:00 AM", "type": "event", "location": "Bucksport Field 1", "team_id": None, "coach_id": None, "notes": "Season opener - all teams"},
                {"id": 2, "title": "Majors Practice", "date": "2025-04-07", "time": "5:30 PM", "type": "practice", "location": "Bucksport Field 1", "team_id": 1, "coach_id": 1, "notes": ""},
                {"id": 3, "title": "Minors Practice", "date": "2025-04-08", "time": "5:30 PM", "type": "practice", "location": "Bucksport Field 2", "team_id": 2, "coach_id": 1, "notes": ""},
                {"id": 4, "title": "Majors vs Ellsworth", "date": "2025-04-12", "time": "1:00 PM", "type": "game", "location": "Bucksport Field 1", "team_id": 1, "coach_id": 1, "notes": "Home game"},
                {"id": 5, "title": "Tee Ball Practice", "date": "2025-04-09", "time": "5:00 PM", "type": "practice", "location": "Bucksport Field 2", "team_id": 3, "coach_id": 1, "notes": ""},
            ]
        
        return [
            {
                "id": e.id,
                "title": e.title,
                "date": e.date,
                "time": e.time,
                "type": e.event_type,
                "location": e.location,
                "team_id": e.team_id,
                "coach_id": e.coach_id,
                "notes": e.notes
            }
            for e in events
        ]
    except Exception as e:
        logger.error(f"Error fetching schedule: {e}")
        # Return sample data on error
        return [
            {"id": 1, "title": "Opening Day", "date": "2025-04-05", "time": "10:00 AM", "type": "event", "location": "Bucksport Field 1", "team_id": None, "coach_id": None, "notes": "Season opener - all teams"},
        ]

@app.get("/api/locations")
def get_locations(session: Session = Depends(get_session)):
    """Get all locations from database."""
    statement = select(Location)
    locations = session.exec(statement).all()
    
    # If no locations in DB, return default list
    if not locations:
        return [
            "Bucksport Field 1",
            "Bucksport Field 2", 
            "Bucksport Softball Field",
            "Miles Lane Complex",
            "Away - Ellsworth",
            "Away - Brewer",
            "Away - Bangor"
        ]
    
    return [loc.name for loc in locations]

class EventRequest(BaseModel):
    title: str
    date: str
    time: str
    location: str
    team_id: str | None = None
    coach_id: str | None = None
    type: str
    notes: str | None = None

@app.post("/api/schedule/request")
def request_new_event(request: EventRequest, session: Session = Depends(get_session)):
    """Create a new scheduled event."""
    event = ScheduleEvent(
        title=request.title,
        date=request.date,
        time=request.time,
        location=request.location,
        event_type=request.type,
        team_id=int(request.team_id) if request.team_id and request.team_id.isdigit() else None,
        coach_id=int(request.coach_id) if request.coach_id and request.coach_id.isdigit() else None,
        notes=request.notes
    )
    session.add(event)
    session.commit()
    session.refresh(event)
    
    logger.info(f"Created new event: {event.title} on {event.date}")
    return {"status": "success", "message": "Event created successfully.", "id": event.id}

@app.put("/api/schedule/{event_id}")
def update_event(event_id: int, request: EventRequest, session: Session = Depends(get_session)):
    """Update a scheduled event."""
    event = session.get(ScheduleEvent, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    event.title = request.title
    event.date = request.date
    event.time = request.time
    event.location = request.location
    event.event_type = request.type
    event.team_id = int(request.team_id) if request.team_id and request.team_id.isdigit() else None
    event.coach_id = int(request.coach_id) if request.coach_id and request.coach_id.isdigit() else None
    event.notes = request.notes
    
    session.add(event)
    session.commit()
    session.refresh(event)
    
    logger.info(f"Updated event: {event.title} (ID: {event_id})")
    return {"status": "success", "message": "Event updated successfully."}

@app.delete("/api/schedule/{event_id}")
def delete_event(event_id: int, session: Session = Depends(get_session)):
    """Delete a scheduled event."""
    event = session.get(ScheduleEvent, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    session.delete(event)
    session.commit()
    logger.info(f"Deleted event: {event.title} (ID: {event_id})")
    return {"status": "success", "message": "Event deleted successfully."}

# ----------------- Inventory endpoints -----------------
@app.get("/api/inventory")
def get_inventory(session: Session = Depends(get_session)):
    """Get all inventory items."""
    statement = select(InventoryItem)
    items = session.exec(statement).all()
    return [
        {
            "id": item.id,
            "name": item.item_name,  # Frontend expects 'name'
            "category": item.category,
            "division": item.division,  # Baseball, Softball, or Shared
            "size": item.size,
            "team": {"id": None, "name": item.team} if item.team else None,
            "assigned_coach": {"id": None, "name": item.assigned_coach} if item.assigned_coach else None,
            "quantity": item.quantity,
            "status": item.status.lower().replace(" ", "-"),  # Convert to lowercase with dashes for badge styling
            "notes": item.notes,
            "last_updated": item.last_updated.isoformat() if item.last_updated else None
        }
        for item in items
    ]

@app.get("/api/inventory/summary")
def get_inventory_summary(session: Session = Depends(get_session)):
    """Get inventory summary statistics."""
    statement = select(InventoryItem)
    items = session.exec(statement).all()
    
    total_quantity = sum(item.quantity for item in items)
    available = sum(item.quantity for item in items if item.status == "Available")
    checked_out = sum(item.quantity for item in items if item.status == "Checked Out")
    needs_repair = sum(item.quantity for item in items if item.status == "Needs Repair")
    
    return {
        "total_items": total_quantity,
        "available": available,
        "checked_out": checked_out,
        "needs_repair": needs_repair
    }

@app.get("/api/inventory/categories")
def get_inventory_categories():
    # Return standard equipment categories
    return ["jersey", "pants", "hat", "cleats", "bat", "ball", "glove", "helmet", "other"]

@app.get("/api/inventory/statuses")
def get_inventory_statuses():
    # Return standard inventory statuses
    return ["Available", "Checked Out", "Needs Repair", "Retired"]

@app.post("/api/inventory", status_code=status.HTTP_201_CREATED)
def create_inventory_item(item_data: dict, session: Session = Depends(get_session)):
    """Create a new inventory item."""
    from datetime import datetime
    
    # Determine division based on category and name
    division = "Shared"
    item_name = item_data.get("name", "").lower()
    category = item_data.get("category", "").lower()
    
    if category in ["jersey", "pants"] and ("girl" in item_name or "women" in item_name or "softball" in item_name):
        division = "Softball"
    elif category in ["jersey", "pants", "bat"] and ("baseball" in item_name or "boy" in item_name or "men" in item_name):
        division = "Baseball"
    
    new_item = InventoryItem(
        item_name=item_data.get("name"),
        category=item_data.get("category"),
        division=division,
        size=item_data.get("size"),
        team=item_data.get("team"),
        assigned_coach=item_data.get("assigned_coach"),
        quantity=item_data.get("quantity", 1),
        status=item_data.get("status", "in-stock"),
        notes=item_data.get("notes", ""),
        last_updated=datetime.now()
    )
    
    session.add(new_item)
    session.commit()
    session.refresh(new_item)
    
    return {
        "id": new_item.id,
        "name": new_item.item_name,
        "category": new_item.category,
        "division": new_item.division,
        "size": new_item.size,
        "team": new_item.team,
        "assigned_coach": new_item.assigned_coach,
        "quantity": new_item.quantity,
        "status": new_item.status,
        "notes": new_item.notes,
        "last_updated": new_item.last_updated.isoformat()
    }

@app.put("/api/inventory/{item_id}")
def update_inventory_item(item_id: int, item_data: dict, session: Session = Depends(get_session)):
    """Update an inventory item."""
    item = session.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Update fields
    if "name" in item_data:
        item.item_name = item_data["name"]
    if "category" in item_data:
        item.category = item_data["category"]
    if "division" in item_data:
        item.division = item_data["division"]
    if "size" in item_data:
        item.size = item_data["size"]
    if "team" in item_data:
        item.team = item_data["team"]
    if "assigned_coach" in item_data:
        item.assigned_coach = item_data["assigned_coach"]
    if "quantity" in item_data:
        item.quantity = item_data["quantity"]
    if "status" in item_data:
        item.status = item_data["status"]
    if "notes" in item_data:
        item.notes = item_data["notes"]
    
    item.last_updated = datetime.utcnow()
    session.add(item)
    session.commit()
    session.refresh(item)
    
    logger.info(f"Updated inventory item: {item.item_name} (ID: {item_id})")
    return {
        "id": item.id,
        "name": item.item_name,
        "category": item.category,
        "division": item.division,
        "size": item.size,
        "team": item.team,
        "assigned_coach": item.assigned_coach,
        "quantity": item.quantity,
        "status": item.status,
        "notes": item.notes,
        "last_updated": item.last_updated.isoformat()
    }

@app.delete("/api/inventory/{item_id}")
def delete_inventory_item(item_id: int, session: Session = Depends(get_session)):
    """Delete an inventory item."""
    item = session.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    item_name = item.item_name
    session.delete(item)
    session.commit()
    
    logger.info(f"Deleted inventory item: {item_name} (ID: {item_id})")
    return {"status": "success", "message": "Item deleted successfully"}

# ----------------- Activity Log endpoints -----------------
@app.get("/api/activity-logs")
def get_activity_logs(
    page: Optional[str] = None,
    limit: int = 1000,
    days: int = 30,
    session: Session = Depends(get_session)
):
    """Get activity logs, optionally filtered by page. Shows last 30 days by default."""
    from models import ActivityLog
    from datetime import datetime, timedelta
    
    statement = select(ActivityLog).order_by(ActivityLog.timestamp.desc())
    
    # Filter by date - last 30 days by default
    cutoff_date = datetime.utcnow() - timedelta(days=days)
    statement = statement.where(ActivityLog.timestamp >= cutoff_date)
    
    if page:
        statement = statement.where(ActivityLog.page == page)
    
    statement = statement.limit(limit)
    logs = session.exec(statement).all()
    
    return [
        {
            "id": log.id,
            "timestamp": log.timestamp.isoformat(),
            "action": log.action,
            "details": log.details,
            "user": log.user,
            "page": log.page,
            "item_id": log.item_id
        }
        for log in logs
    ]

@app.post("/api/activity-logs")
def create_activity_log(log_data: dict, session: Session = Depends(get_session)):
    """Create a new activity log entry."""
    from models import ActivityLog
    
    log = ActivityLog(
        action=log_data.get("action", "Unknown Action"),
        details=log_data.get("details", ""),
        user=log_data.get("user", "Unknown User"),
        page=log_data.get("page", "unknown"),
        item_id=log_data.get("item_id")
    )
    
    session.add(log)
    session.commit()
    session.refresh(log)
    
    logger.info(f"Activity logged: {log.action} by {log.user} on {log.page}")
    return {"status": "success", "id": log.id}

# ----------------- Donation endpoints -----------------
@app.get("/api/donations")
def get_donations(
    donation_type: Optional[str] = None,
    division: Optional[str] = None,
    session: Session = Depends(get_session)
):
    """Get all donations, optionally filtered by type or division."""
    from models import Donation
    
    statement = select(Donation).order_by(Donation.date.desc())
    
    if donation_type:
        statement = statement.where(Donation.donation_type == donation_type)
    
    if division:
        statement = statement.where(Donation.division == division)
    
    donations = session.exec(statement).all()
    
    return [
        {
            "id": d.id,
            "name": d.name,
            "amount": d.amount,
            "type": d.donation_type,
            "date": d.date.isoformat(),
            "division": d.division,
            "contact_person": d.contact_person,
            "phone": d.phone,
            "email": d.email,
            "address": d.address,
            "notes": d.notes
        }
        for d in donations
    ]

@app.post("/api/donations")
def create_donation(donation_data: dict, session: Session = Depends(get_session)):
    """Create a new donation entry."""
    from models import Donation
    from datetime import datetime
    
    # Parse date
    date_str = donation_data.get("date")
    if isinstance(date_str, str):
        donation_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        donation_date = date_str
    
    donation = Donation(
        name=donation_data.get("name"),
        amount=float(donation_data.get("amount")),
        donation_type=donation_data.get("type", "Donation"),
        date=donation_date,
        division=donation_data.get("division"),
        contact_person=donation_data.get("contact_person"),
        phone=donation_data.get("phone"),
        email=donation_data.get("email"),
        address=donation_data.get("address"),
        notes=donation_data.get("notes")
    )
    
    session.add(donation)
    session.commit()
    session.refresh(donation)
    
    logger.info(f"Created donation: {donation.name} - ${donation.amount}")
    
    return {
        "id": donation.id,
        "name": donation.name,
        "amount": donation.amount,
        "type": donation.donation_type,
        "date": donation.date.isoformat(),
        "division": donation.division,
        "contact_person": donation.contact_person,
        "phone": donation.phone,
        "email": donation.email,
        "address": donation.address,
        "notes": donation.notes
    }

@app.delete("/api/donations/{donation_id}")
def delete_donation(donation_id: int, session: Session = Depends(get_session)):
    """Delete a donation by ID."""
    from models import Donation
    
    donation = session.get(Donation, donation_id)
    if not donation:
        raise HTTPException(status_code=404, detail="Donation not found")
    
    donation_name = donation.name
    donation_amount = donation.amount
    
    session.delete(donation)
    session.commit()
    
    logger.info(f"Deleted donation: {donation_name} - ${donation_amount} (ID: {donation_id})")
    return {"status": "success", "message": "Donation deleted successfully"}


@app.get("/api/sponsorship-sheets")
def list_sponsorship_sheets(session: Session = Depends(get_session)):
    from models import SponsorshipSheetMeta
    metas = session.exec(select(SponsorshipSheetMeta)).all()
    
    # Define explicit order to match Excel import order
    sheet_order = [
        "Master Sponsor List",
        "Softball Banners - Current",
        "Softball Banners - Team Sponsor",
        "Baseball Banners - Current",
    ]
    
    # Sort metas by the defined order
    def get_order(meta):
        try:
            return sheet_order.index(meta.sheet_name)
        except ValueError:
            return len(sheet_order)  # Put unknown sheets at the end
    
    sorted_metas = sorted(metas, key=get_order)
    
    return [
        {
            "sheet_name": m.sheet_name,
            "columns": m.columns,
            "updated_at": m.updated_at.isoformat() if m.updated_at else None,
        }
        for m in sorted_metas
    ]


@app.get("/api/sponsorship-sheets/{sheet_name}")
def get_sponsorship_sheet(sheet_name: str, session: Session = Depends(get_session)):
    from models import SponsorshipSheetMeta, SponsorshipSheetRow

    meta = session.get(SponsorshipSheetMeta, sheet_name)
    if not meta:
        raise HTTPException(status_code=404, detail="Sheet not found")

    rows = session.exec(
        select(SponsorshipSheetRow)
        .where(SponsorshipSheetRow.sheet_name == sheet_name)
        .order_by(SponsorshipSheetRow.row_index.asc())
    ).all()

    return {
        "sheet_name": meta.sheet_name,
        "columns": meta.columns,
        "rows": [
            {
                "id": r.id,
                "row_index": r.row_index,
                "data": r.data,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
            for r in rows
        ],
        "updated_at": meta.updated_at.isoformat() if meta.updated_at else None,
    }


@app.post("/api/sponsorship-sheets/{sheet_name}/rows", status_code=status.HTTP_201_CREATED)
def create_sponsorship_sheet_row(
    sheet_name: str,
    payload: dict,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_fundraising_editor)
):
    from models import SponsorshipSheetMeta, SponsorshipSheetRow
    from datetime import datetime

    meta = session.get(SponsorshipSheetMeta, sheet_name)
    if not meta:
        raise HTTPException(status_code=404, detail="Sheet not found")

    existing_max = session.exec(
        select(SponsorshipSheetRow.row_index)
        .where(SponsorshipSheetRow.sheet_name == sheet_name)
        .order_by(SponsorshipSheetRow.row_index.desc())
    ).first()

    next_row_index = (existing_max or 1) + 1
    row_data = payload.get("data") if isinstance(payload, dict) else None
    if row_data is None or not isinstance(row_data, dict):
        row_data = {}

    new_row = SponsorshipSheetRow(
        sheet_name=sheet_name,
        row_index=next_row_index,
        data=row_data,
        updated_at=datetime.utcnow(),
    )
    session.add(new_row)
    session.commit()
    session.refresh(new_row)

    meta.updated_at = datetime.utcnow()
    session.add(meta)
    session.commit()

    return {
        "id": new_row.id,
        "sheet_name": new_row.sheet_name,
        "row_index": new_row.row_index,
        "data": new_row.data,
        "updated_at": new_row.updated_at.isoformat() if new_row.updated_at else None,
    }


@app.put("/api/sponsorship-sheets/{sheet_name}/rows/{row_index}")
def upsert_sponsorship_sheet_row(
    sheet_name: str,
    row_index: int,
    payload: dict,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_fundraising_editor)
):
    from models import SponsorshipSheetMeta, SponsorshipSheetRow
    from datetime import datetime

    meta = session.get(SponsorshipSheetMeta, sheet_name)
    if not meta:
        raise HTTPException(status_code=404, detail="Sheet not found")

    row = session.exec(
        select(SponsorshipSheetRow)
        .where(SponsorshipSheetRow.sheet_name == sheet_name)
        .where(SponsorshipSheetRow.row_index == row_index)
    ).first()

    row_data = payload.get("data") if isinstance(payload, dict) else None
    if row_data is None or not isinstance(row_data, dict):
        raise HTTPException(status_code=400, detail="Payload must include data object")

    if row:
        row.data = row_data
        row.updated_at = datetime.utcnow()
        session.add(row)
        session.commit()
        session.refresh(row)
    else:
        row = SponsorshipSheetRow(
            sheet_name=sheet_name,
            row_index=row_index,
            data=row_data,
            updated_at=datetime.utcnow(),
        )
        session.add(row)
        session.commit()
        session.refresh(row)

    meta.updated_at = datetime.utcnow()
    session.add(meta)
    session.commit()

    return {
        "id": row.id,
        "sheet_name": row.sheet_name,
        "row_index": row.row_index,
        "data": row.data,
        "updated_at": row.updated_at.isoformat() if row.updated_at else None,
    }


@app.delete("/api/sponsorship-sheets/{sheet_name}/rows/{row_index}")
def delete_sponsorship_sheet_row(
    sheet_name: str,
    row_index: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_fundraising_editor)
):
    from models import SponsorshipSheetMeta, SponsorshipSheetRow
    from datetime import datetime

    meta = session.get(SponsorshipSheetMeta, sheet_name)
    if not meta:
        raise HTTPException(status_code=404, detail="Sheet not found")

    row = session.exec(
        select(SponsorshipSheetRow)
        .where(SponsorshipSheetRow.sheet_name == sheet_name)
        .where(SponsorshipSheetRow.row_index == row_index)
    ).first()

    if not row:
        raise HTTPException(status_code=404, detail="Row not found")

    session.delete(row)
    session.commit()

    meta.updated_at = datetime.utcnow()
    session.add(meta)
    session.commit()

    return {"status": "success"}


@app.post("/api/sponsorship-sheets/{sheet_name}/columns")
def add_column_to_sheet(
    sheet_name: str,
    payload: dict,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_fundraising_editor)
):
    from models import SponsorshipSheetMeta, SponsorshipSheetRow
    from datetime import datetime

    meta = session.get(SponsorshipSheetMeta, sheet_name)
    if not meta:
        raise HTTPException(status_code=404, detail="Sheet not found")

    column_name = payload.get("column_name")
    if not column_name or not isinstance(column_name, str):
        raise HTTPException(status_code=400, detail="column_name is required")

    # Check if column already exists
    if column_name in meta.columns:
        raise HTTPException(status_code=400, detail="Column already exists")

    # Add column to metadata
    meta.columns.append(column_name)
    meta.updated_at = datetime.utcnow()
    
    # Update all existing rows to include the new column with empty value
    rows = session.exec(
        select(SponsorshipSheetRow)
        .where(SponsorshipSheetRow.sheet_name == sheet_name)
    ).all()
    
    for row in rows:
        if column_name not in row.data:
            row.data[column_name] = ""
            row.updated_at = datetime.utcnow()
            session.add(row)
    
    session.add(meta)
    session.commit()
    session.refresh(meta)

    return {
        "sheet_name": meta.sheet_name,
        "columns": meta.columns,
        "updated_at": meta.updated_at.isoformat() if meta.updated_at else None,
    }


# TEMPORARY ADMIN ENDPOINT - Remove after first use
@app.post("/api/admin/setup-donations")
def setup_donations_from_spreadsheet(session: Session = Depends(get_session)):
    """
    ONE-TIME SETUP: Import donation data from spreadsheet.
    This endpoint should be called once to populate the database, then removed.
    """
    from models import Donation
    from datetime import date
    import pandas as pd
    
    try:
        # Check if data already exists
        existing = session.exec(select(Donation)).first()
        if existing:
            return {
                "status": "already_setup",
                "message": "Donation data already exists in database",
                "count": len(session.exec(select(Donation)).all())
            }
        
        # Load Excel file
        xl = pd.ExcelFile('/opt/render/project/src/Softball AND Baseball Banner & Sponsorship Log.xlsx')
        total_imported = 0
        
        # Process Master Sponsor List
        df_master = pd.read_excel(xl, sheet_name='Master Sponsor List')
        for _, row in df_master.iterrows():
            company_name = row.get('Company Name')
            if pd.isna(company_name) or company_name == '':
                continue
            
            for year in ['2025', '2024', '2023', '2022', '2021', '2020']:
                amount = row.get(year)
                if pd.notna(amount):
                    try:
                        amount_float = float(amount)
                        if amount_float > 0:
                            donation = Donation(
                                name=str(company_name),
                                amount=amount_float,
                                donation_type='Sponsorship',
                                date=date(int(year), 1, 1),
                                division=row.get('Division') if pd.notna(row.get('Division')) else None,
                                contact_person=row.get('Contact Person') if pd.notna(row.get('Contact Person')) else None,
                                phone=row.get('Phone') if pd.notna(row.get('Phone')) else None,
                                email=row.get('Email') if pd.notna(row.get('Email')) else None,
                                address=row.get('Address') if pd.notna(row.get('Address')) else None,
                                notes=f"{row.get('Sponsor Type', '')} - {row.get('Notes', '')}" if pd.notna(row.get('Notes')) else row.get('Sponsor Type', '')
                            )
                            session.add(donation)
                            total_imported += 1
                    except (ValueError, TypeError):
                        continue
        
        session.commit()
        
        return {
            "status": "success",
            "message": f"Successfully imported {total_imported} donation records",
            "total_imported": total_imported
        }
        
    except Exception as e:
        logger.error(f"Error importing donations: {str(e)}")
        return {
            "status": "error",
            "message": f"Failed to import donations: {str(e)}"
        }


@app.get("/api/user/permissions")
async def get_user_permissions(current_user: User = Depends(get_current_user)):
    """Get current user's permissions for fundraising."""
    return {
        "email": current_user.email,
        "role": current_user.role,
        "can_edit_fundraising": can_edit_fundraising(current_user),
        "first_name": current_user.first_name,
        "last_name": current_user.last_name
    }


@app.get("/api/sponsorship-sheets/export/excel")
async def export_sponsorship_sheets_to_excel(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Export all sponsorship sheets to Excel file."""
    from models import SponsorshipSheetMeta, SponsorshipSheetRow
    
    # Define sheet order
    sheet_order = [
        "Master Sponsor List",
        "Softball Banners - Current",
        "Softball Banners - Team Sponsor",
        "Baseball Banners - Current",
    ]
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for sheet_name in sheet_order:
        meta = session.get(SponsorshipSheetMeta, sheet_name)
        if not meta:
            continue
            
        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(meta.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Get rows
        rows = session.exec(
            select(SponsorshipSheetRow)
            .where(SponsorshipSheetRow.sheet_name == sheet_name)
            .order_by(SponsorshipSheetRow.row_index.asc())
        ).all()
        
        # Write data rows
        for row_obj in rows:
            excel_row = row_obj.row_index + 1  # +1 because header is row 1
            for col_idx, col_name in enumerate(meta.columns, start=1):
                cell_value = row_obj.data.get(col_name, "")
                ws.cell(row=excel_row, column=col_idx, value=cell_value)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Sponsorship_Log_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Mount the static directory to serve frontend files. This should be last.
app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")
