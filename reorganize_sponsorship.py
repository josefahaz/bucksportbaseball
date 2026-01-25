import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil

# Create backup
shutil.copy('Softball AND Baseball Banner & Sponsorship Log.xlsx', 
            f'Softball AND Baseball Banner & Sponsorship Log_BACKUP_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

# Load the workbook
wb = openpyxl.load_workbook('Softball AND Baseball Banner & Sponsorship Log.xlsx')

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_row_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def clean_value(val):
    """Clean cell values."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return val

# Create new consolidated sheet
if "Master Sponsor List" in wb.sheetnames:
    del wb["Master Sponsor List"]
master_ws = wb.create_sheet("Master Sponsor List", 0)

# Set up master sheet headers
headers = [
    "Sponsor Type", "Company Name", "Contact Person", "Phone", "Email", 
    "Address", "Division", "Status", "2025", "2024", "2023", "2022", "2021", "2020", "Notes"
]

for col_idx, header in enumerate(headers, 1):
    cell = master_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Collect all sponsor data
all_sponsors = []

# Process Softball Banners - Current (Field Sponsors)
print("Processing Softball Banners - Current...")
softball_current_ws = wb["Softball Banners - Current"]
for row_idx in range(2, softball_current_ws.max_row + 1):
    status = clean_value(softball_current_ws.cell(row=row_idx, column=1).value)
    company = clean_value(softball_current_ws.cell(row=row_idx, column=2).value)
    
    if not company or company == "":
        continue
    
    contact = clean_value(softball_current_ws.cell(row=row_idx, column=4).value)
    phone = clean_value(softball_current_ws.cell(row=row_idx, column=5).value)
    notes = clean_value(softball_current_ws.cell(row=row_idx, column=6).value)
    
    # Extract email from phone/contact field if present
    email = ""
    if "@" in str(phone):
        email = phone
        phone = ""
    
    sponsor_data = {
        "type": "Field Banner",
        "company": company,
        "contact": contact,
        "phone": phone,
        "email": email,
        "address": "",
        "division": "Softball",
        "status": status,
        "2025": clean_value(softball_current_ws.cell(row=row_idx, column=13).value),
        "2024": "",
        "2023": "",
        "2022": "",
        "2021": "",
        "2020": "",
        "notes": notes
    }
    all_sponsors.append(sponsor_data)

# Process Softball Banners - Team Sponsors
print("Processing Softball Banners - Team Sponsors...")
softball_ws = wb["Softball Banners - Team Sponsor"]
for row_idx in range(2, softball_ws.max_row + 1):
    company = clean_value(softball_ws.cell(row=row_idx, column=1).value)
    
    if not company or company == "":
        continue
    
    contact = clean_value(softball_ws.cell(row=row_idx, column=2).value)
    phone = clean_value(softball_ws.cell(row=row_idx, column=3).value)
    email = clean_value(softball_ws.cell(row=row_idx, column=4).value)
    address = clean_value(softball_ws.cell(row=row_idx, column=5).value)
    notes = clean_value(softball_ws.cell(row=row_idx, column=6).value)
    
    sponsor_data = {
        "type": "Team Sponsor",
        "company": company,
        "contact": contact,
        "phone": phone,
        "email": email,
        "address": address,
        "division": "Softball",
        "status": "Active" if softball_ws.cell(row=row_idx, column=7).value else "",
        "2025": clean_value(softball_ws.cell(row=row_idx, column=7).value),
        "2024": clean_value(softball_ws.cell(row=row_idx, column=8).value),
        "2023": clean_value(softball_ws.cell(row=row_idx, column=9).value),
        "2022": clean_value(softball_ws.cell(row=row_idx, column=10).value),
        "2021": clean_value(softball_ws.cell(row=row_idx, column=11).value),
        "2020": clean_value(softball_ws.cell(row=row_idx, column=12).value),
        "notes": notes
    }
    all_sponsors.append(sponsor_data)

# Process Baseball Banners - Current
print("Processing Baseball Banners - Current...")
baseball_ws = wb["Baseball Banners - Current"]
for row_idx in range(2, baseball_ws.max_row + 1):
    company = clean_value(baseball_ws.cell(row=row_idx, column=2).value)
    
    if not company or company == "":
        continue
    
    status = clean_value(baseball_ws.cell(row=row_idx, column=1).value)
    contact = clean_value(baseball_ws.cell(row=row_idx, column=4).value)
    address = clean_value(baseball_ws.cell(row=row_idx, column=5).value)
    notes = clean_value(baseball_ws.cell(row=row_idx, column=6).value)
    
    # Extract email and phone from address field if present
    email = ""
    phone = ""
    if "@" in str(address):
        parts = str(address).split()
        for part in parts:
            if "@" in part:
                email = part
    
    sponsor_data = {
        "type": "Field Banner",
        "company": company,
        "contact": contact,
        "phone": phone,
        "email": email,
        "address": address,
        "division": "Baseball",
        "status": status,
        "2025": clean_value(baseball_ws.cell(row=row_idx, column=8).value),
        "2024": "",
        "2023": "",
        "2022": "",
        "2021": "",
        "2020": "",
        "notes": notes
    }
    all_sponsors.append(sponsor_data)

# Sort sponsors by company name
all_sponsors.sort(key=lambda x: x["company"].lower() if x["company"] else "")

# Write to master sheet
print(f"\nWriting {len(all_sponsors)} sponsors to Master Sponsor List...")
for idx, sponsor in enumerate(all_sponsors, 2):
    row_data = [
        sponsor["type"],
        sponsor["company"],
        sponsor["contact"],
        sponsor["phone"],
        sponsor["email"],
        sponsor["address"],
        sponsor["division"],
        sponsor["status"],
        sponsor["2025"],
        sponsor["2024"],
        sponsor["2023"],
        sponsor["2022"],
        sponsor["2021"],
        sponsor["2020"],
        sponsor["notes"]
    ]
    
    for col_idx, value in enumerate(row_data, 1):
        cell = master_ws.cell(row=idx, column=col_idx, value=value)
        cell.border = border
        
        # Alternate row colors
        if idx % 2 == 0:
            cell.fill = alt_row_fill
        
        # Alignment
        if col_idx in [1, 7, 8]:  # Type, Division, Status
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Set column widths
column_widths = {
    'A': 15,  # Type
    'B': 30,  # Company
    'C': 20,  # Contact
    'D': 15,  # Phone
    'E': 25,  # Email
    'F': 35,  # Address
    'G': 12,  # Division
    'H': 15,  # Status
    'I': 10,  # 2025
    'J': 10,  # 2024
    'K': 10,  # 2023
    'L': 10,  # 2022
    'M': 10,  # 2021
    'N': 10,  # 2020
    'O': 40   # Notes
}

for col, width in column_widths.items():
    master_ws.column_dimensions[col].width = width

# Freeze top row
master_ws.freeze_panes = 'A2'

# Add summary at the bottom
summary_row = len(all_sponsors) + 3
master_ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)

summary_row += 1
master_ws.cell(row=summary_row, column=1, value="Total Sponsors:")
master_ws.cell(row=summary_row, column=2, value=len(all_sponsors))

summary_row += 1
baseball_count = sum(1 for s in all_sponsors if s["division"] == "Baseball")
master_ws.cell(row=summary_row, column=1, value="Baseball Sponsors:")
master_ws.cell(row=summary_row, column=2, value=baseball_count)

summary_row += 1
softball_count = sum(1 for s in all_sponsors if s["division"] == "Softball")
master_ws.cell(row=summary_row, column=1, value="Softball Sponsors:")
master_ws.cell(row=summary_row, column=2, value=softball_count)

summary_row += 1
active_2025 = sum(1 for s in all_sponsors if s["2025"])
master_ws.cell(row=summary_row, column=1, value="Active 2025:")
master_ws.cell(row=summary_row, column=2, value=active_2025)

# Save the workbook
wb.save('Softball AND Baseball Banner & Sponsorship Log.xlsx')
print("\n✓ Successfully reorganized sponsorship file!")
print(f"✓ Created 'Master Sponsor List' sheet with {len(all_sponsors)} sponsors")
print(f"✓ Baseball: {baseball_count} | Softball: {softball_count} | Active 2025: {active_2025}")
print("\nBackup created with timestamp.")
