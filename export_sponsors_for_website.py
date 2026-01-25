import openpyxl
import csv

# Load the workbook
wb = openpyxl.load_workbook('Softball AND Baseball Banner & Sponsorship Log.xlsx')
master_ws = wb["Master Sponsor List"]

# Prepare data for website - only active 2025 sponsors
website_sponsors = []

print("Extracting active 2025 sponsors for website...")

for row_idx in range(2, master_ws.max_row + 1):
    sponsor_type = master_ws.cell(row=row_idx, column=1).value
    company = master_ws.cell(row=row_idx, column=2).value
    division = master_ws.cell(row=row_idx, column=7).value
    amount_2025 = master_ws.cell(row=row_idx, column=9).value
    
    # Only include sponsors with 2025 activity
    if not company or not amount_2025:
        continue
    
    # Determine sponsorship level based on amount
    level = "Supporter"
    if amount_2025:
        try:
            amount = float(amount_2025) if isinstance(amount_2025, (int, float)) else 0
            if amount >= 1000:
                level = "Gold Sponsor"
            elif amount >= 500:
                level = "Silver Sponsor"
            elif amount >= 250:
                level = "Bronze Sponsor"
        except:
            level = "Supporter"
    
    website_sponsors.append({
        "company": company,
        "type": sponsor_type,
        "division": division,
        "level": level,
        "amount": amount_2025
    })

# Sort by amount (highest first), then alphabetically
website_sponsors.sort(key=lambda x: (-float(x["amount"]) if isinstance(x["amount"], (int, float)) else 0, x["company"].lower()))

# Export to CSV for website
csv_filename = "sponsors_for_website.csv"
with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=["company", "type", "division", "level", "amount"])
    writer.writeheader()
    writer.writerows(website_sponsors)

print(f"\n✓ Exported {len(website_sponsors)} active 2025 sponsors to {csv_filename}")

# Create a summary by level
levels = {}
for sponsor in website_sponsors:
    level = sponsor["level"]
    levels[level] = levels.get(level, 0) + 1

print("\nSponsorship Levels:")
for level in ["Gold Sponsor", "Silver Sponsor", "Bronze Sponsor", "Supporter"]:
    if level in levels:
        print(f"  {level}: {levels[level]}")

# Create a summary by division
divisions = {}
for sponsor in website_sponsors:
    div = sponsor["division"]
    divisions[div] = divisions.get(div, 0) + 1

print("\nBy Division:")
for div, count in sorted(divisions.items()):
    print(f"  {div}: {count}")

wb.close()
print(f"\n✓ Ready for website integration!")
